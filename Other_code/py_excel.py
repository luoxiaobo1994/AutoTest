# -*- coding:utf-8 -*-
# Author: luoxiaobo
# TIME: 2021-08-12 10:05

import openpyxl,pandas

fp = r"D:\data\xlsx_test.xlsx"

def open_way():
    workbook = openpyxl.load_workbook(fp)
    sheet = workbook.sheetnames  # 属性,直接获取工作表的表名  -->['Sheet1', 'Sheet2', 'Sheet3']
    sh1 = workbook['Sheet1']  # 可以用表名称引入
    # sh1.append(['11A', '11B'])
    row2 = [item.value for item in list(sh1.rows)[1]]  # 第2行的值,从0开始,需要转列表.
    col1 = [item.value for item in list(sh1.columns)[0]]  # 第1列的值.
    cell_2_3 = sh1.cell(row=2,column=3).value  # 指定单元格的值.单个元素
    max_row = sh1.max_row
    max_col = sh1.max_column
    # workbook.save(fp)
    print(row2,col1,cell_2_3,max_row,max_col)

def pd_way():
    data = pandas.read_excel(fp,header=None)
    print(data)

if __name__ == '__main__':
    pd_way()
    # open_way()