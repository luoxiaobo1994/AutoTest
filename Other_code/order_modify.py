# -*- coding:utf-8 -*-
# Author: luoxiaobo
# TIME: 2021/9/10 10:32

import openpyxl

file = r"E:\工作\项目\订单\xlsx订单\多商品多订单OrderPicking - 副本.xlsx"

wb = openpyxl.load_workbook(file)  # 加载表格
sheet1 = wb['Sheet1']  # 读取第一个工作表
rows = sheet1.max_row  # 最大行数。
for i in range(3,rows):  # 序号是从1开始的，前两行是注释+表头。数据从3开始。
    try:
        old = sheet1.cell(row=i, column=1).value
        print(sheet1.cell(row=i, column=1).value)
        sheet1.cell(row=i, column=1).value = str(int(old) + 100)
    except:
        print("有非数字的订单ID")
wb.save(filename=file)
# for row in range(2)
